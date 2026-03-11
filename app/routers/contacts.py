import csv
import io
from uuid import UUID

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from .. import crud, models, schemas
from ..deps import get_current_user_id, get_db

router = APIRouter(prefix="/tenants/{tenant_id}/contacts", tags=["contacts"])

COLUMNS = ["מקצוע", "שם המשרד", "איש קשר", "טלפון משרד", "טלפון נייד", "אימייל", "פרטים"]

SAMPLE_ROWS = [
    ["אדריכל", "משרד כהן אדריכלים", "דוד כהן", "03-1234567", "050-1234567", "david@cohen-arch.co.il", "מתמחה בבנייה למגורים"],
    ["עורך דין", "מזרחי ושות", "רונית מזרחי", "03-7654321", "052-7654321", "ronit@mazrahi.law", "נדל\"ן ומקרקעין"],
    ["קבלן ראשי", "בן דוד בנייה", "יוסי בן דוד", "04-1112233", "054-1112233", "yosi@bendavid.co.il", "ניסיון 20 שנה"],
    ["מהנדס קונסטרוקציה", "הנדסה ופתרונות", "מיכל לוי", "08-9998877", "053-9998877", "michal@eng.co.il", ""],
    ["שמאי מקרקעין", "שמאות מקצועית", "אבי גולן", "02-3334455", "055-3334455", "avi@shomaut.co.il", "מוסמך מועצת השמאים"],
]


def _get_contact_or_404(db: Session, tenant_id: UUID, contact_id: UUID) -> models.Contact:
    contact = (
        db.query(models.Contact)
        .filter(
            models.Contact.id == contact_id,
            models.Contact.tenant_id == tenant_id,
            models.Contact.deleted_at.is_(None),
        )
        .first()
    )
    if not contact:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="איש קשר לא נמצא")
    return contact


def _build_xlsx_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "אנשי קשר"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color="011E41", end_color="011E41", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    sample_font = Font(name="Arial", size=10, color="555555")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = sample_font
            cell.alignment = Alignment(horizontal="right")

    col_widths = [22, 24, 20, 16, 16, 32, 28]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/template-xlsx", summary="הורדת תבנית Excel לייבוא אנשי קשר")
def download_template(tenant_id: UUID):
    data = _build_xlsx_template()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''contacts_template.xlsx"},
    )


@router.post("/", response_model=schemas.ContactRead, status_code=status.HTTP_201_CREATED)
def create_contact(
    tenant_id: UUID,
    contact_in: schemas.ContactCreate,
    db: Session = Depends(get_db),
    user_id: str | None = Depends(get_current_user_id),
):
    contact = crud.create_entity(
        db,
        models.Contact,
        contact_in.model_dump(),
        tenant_id=str(tenant_id),
        created_by=user_id,
    )
    db.commit()
    db.refresh(contact)
    return contact


@router.get("/{contact_id}", response_model=schemas.ContactRead)
def read_contact(tenant_id: UUID, contact_id: UUID, db: Session = Depends(get_db)):
    return _get_contact_or_404(db, tenant_id, contact_id)


@router.get("/", response_model=list[schemas.ContactRead])
def list_contacts(tenant_id: UUID, db: Session = Depends(get_db)):
    return (
        db.query(models.Contact)
        .filter(models.Contact.tenant_id == tenant_id, models.Contact.deleted_at.is_(None))
        .all()
    )


@router.put("/{contact_id}", response_model=schemas.ContactRead)
def update_contact(
    tenant_id: UUID,
    contact_id: UUID,
    contact_in: schemas.ContactUpdate,
    db: Session = Depends(get_db),
    changed_by: str | None = Depends(get_current_user_id),
):
    contact = _get_contact_or_404(db, tenant_id, contact_id)
    contact = crud.update_entity(db, contact, contact_in.model_dump(), changed_by=changed_by)
    db.commit()
    db.refresh(contact)
    return contact


@router.post("/import-xlsx", summary="ייבוא אנשי קשר מ-Excel")
async def import_contacts_xlsx(
    tenant_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user_id: str | None = Depends(get_current_user_id),
):
    """מייבא אנשי קשר מקובץ Excel (.xlsx)."""
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="קובץ Excel לא תקין")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0}

    # First row = headers
    headers = [str(h).strip() if h else "" for h in rows[0]]

    def col(row_vals: tuple, col_name: str) -> str:
        try:
            idx = headers.index(col_name)
            val = row_vals[idx]
            return str(val).strip() if val is not None else ""
        except ValueError:
            return ""

    created = []
    for row_vals in rows[1:]:
        name = col(row_vals, "איש קשר")
        if not name:
            continue
        contact = crud.create_entity(
            db,
            models.Contact,
            {
                "name": name,
                "profession": col(row_vals, "מקצוע") or None,
                "office_name": col(row_vals, "שם המשרד") or None,
                "phone": col(row_vals, "טלפון משרד") or None,
                "mobile_phone": col(row_vals, "טלפון נייד") or None,
                "email": col(row_vals, "אימייל") or None,
                "notes": col(row_vals, "פרטים") or None,
            },
            tenant_id=str(tenant_id),
            created_by=user_id,
        )
        created.append(contact)
    db.commit()
    return {"imported": len(created)}


@router.delete("/{contact_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_contact(
    tenant_id: UUID,
    contact_id: UUID,
    db: Session = Depends(get_db),
    changed_by: str | None = Depends(get_current_user_id),
):
    contact = _get_contact_or_404(db, tenant_id, contact_id)
    crud.soft_delete_entity(db, contact, changed_by=changed_by)
    db.commit()
    return None
