"""
יצוא נתונים ל-Excel — משימות ותקציב לפי פרויקט.
"""
import io
from urllib.parse import quote
from uuid import UUID

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from .. import models, schemas
from ..deps import get_db

router = APIRouter(prefix="/tenants/{tenant_id}", tags=["export"])

HEADER_FILL = PatternFill("solid", fgColor="011E41")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GOLD_FILL = PatternFill("solid", fgColor="FCD562")
GOLD_FONT = Font(color="011E41", bold=True)


def _header(ws, row, cols):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="right")


def _fmt_dt(dt):
    return dt.strftime("%d/%m/%Y") if dt else ""


@router.get("/budget/entries", response_model=list[schemas.BudgetEntryRead])
def list_all_budget_entries(tenant_id: UUID, db: Session = Depends(get_db)):
    """כל רשומות התקציב של הטנאנט — קריאה אחת לדשבורד."""
    return (
        db.query(models.BudgetEntry)
        .filter(
            models.BudgetEntry.tenant_id == str(tenant_id),
            models.BudgetEntry.deleted_at.is_(None),
        )
        .all()
    )


@router.get("/projects/{project_id}/export")
def export_project(tenant_id: UUID, project_id: UUID, db: Session = Depends(get_db)):
    project = (
        db.query(models.Project)
        .filter(models.Project.id == project_id, models.Project.tenant_id == tenant_id, models.Project.deleted_at.is_(None))
        .first()
    )
    stages = db.query(models.Stage).filter(models.Stage.project_id == project_id, models.Stage.deleted_at.is_(None)).all()
    tasks = db.query(models.Task).filter(models.Task.project_id == project_id, models.Task.deleted_at.is_(None)).all()
    entries = db.query(models.BudgetEntry).filter(models.BudgetEntry.project_id == project_id, models.BudgetEntry.deleted_at.is_(None)).all()
    users = {str(u.id): u.name for u in db.query(models.User).filter(models.User.tenant_id == tenant_id).all()}
    stage_map = {str(s.id): s.name for s in stages}

    STATUS_HE = {"todo": "לביצוע", "in_progress": "בעבודה", "done": "הושלם", "blocked": "חסום", "review": "לבדיקה"}
    PRIORITY_HE = {"high": "גבוהה", "medium": "בינונית", "low": "נמוכה"}

    wb = Workbook()

    # ---- Sheet 1: משימות ----
    ws1 = wb.active
    ws1.title = "משימות"
    ws1.sheet_view.rightToLeft = True
    _header(ws1, 1, ["קבוצה", "משימה", "סטטוס", "עדיפות", "איש צוות", "התחלה", "סיום"])
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 20
    ws1.column_dimensions["F"].width = 14
    ws1.column_dimensions["G"].width = 14

    for t in sorted(tasks, key=lambda x: x.stage_id):
        ws1.append([
            stage_map.get(str(t.stage_id), ""),
            t.title,
            STATUS_HE.get(t.status, t.status),
            PRIORITY_HE.get(t.priority, t.priority),
            users.get(str(t.assignee_id), "") if t.assignee_id else "",
            _fmt_dt(t.start_date),
            _fmt_dt(t.end_date),
        ])

    # ---- Sheet 2: תקציב ----
    ws2 = wb.create_sheet("תקציב")
    ws2.sheet_view.rightToLeft = True
    _header(ws2, 1, ["קטגוריה", "תיאור", "ספק", "סכום", "סוג", "תאריך"])
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 14

    total_planned = 0
    total_actual = 0
    for e in entries:
        ws2.append([
            e.category,
            e.description,
            e.vendor or "",
            e.amount,
            "מתוכנן" if e.is_planned else "בפועל",
            _fmt_dt(e.entry_date),
        ])
        if e.is_planned:
            total_planned += e.amount
        else:
            total_actual += e.amount

    # סיכום תקציב
    ws2.append([])
    summary_row = ws2.max_row + 1
    ws2.append(["", "סה\"כ מתוכנן", "", total_planned, "", ""])
    ws2.append(["", "סה\"כ בפועל", "", total_actual, "", ""])
    ws2.append(["", "יתרה", "", total_planned - total_actual, "", ""])

    for r in range(summary_row, ws2.max_row + 1):
        for c in range(1, 7):
            cell = ws2.cell(row=r, column=c)
            cell.fill = GOLD_FILL
            cell.font = GOLD_FONT

    # ---- Sheet 3: הצעות מחיר ----
    quotes = db.query(models.Quote).filter(models.Quote.project_id == project_id, models.Quote.deleted_at.is_(None)).all()
    if quotes:
        ws3 = wb.create_sheet("הצעות מחיר")
        ws3.sheet_view.rightToLeft = True
        _header(ws3, 1, ["כותרת", "ספק", "סכום כולל", "סטטוס", "תאריך"])
        ws3.column_dimensions["A"].width = 30
        ws3.column_dimensions["B"].width = 20
        ws3.column_dimensions["C"].width = 16
        ws3.column_dimensions["D"].width = 18
        ws3.column_dimensions["E"].width = 14

        STATUS_HE_Q = {"pending_review": "ממתין לאישור", "approved": "מאושר", "rejected": "נדחה"}
        for q in quotes:
            ws3.append([
                q.title,
                q.vendor or "",
                q.total_amount or 0,
                STATUS_HE_Q.get(q.status, q.status),
                _fmt_dt(q.created_at),
            ])

    # שם הקובץ — RFC 5987 encoding לתמיכה בעברית
    proj_name = (project.name if project else "project").replace(" ", "_")
    filename = f"maslul_{proj_name}.xlsx"
    filename_encoded = quote(filename, safe="")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"maslul_project.xlsx\"; filename*=UTF-8''{filename_encoded}"},
    )


@router.get("/budget/export")
def export_budget(tenant_id: UUID, db: Session = Depends(get_db)):
    """יצוא תקציב כלל-פרויקטי."""
    projects = db.query(models.Project).filter(models.Project.tenant_id == tenant_id, models.Project.deleted_at.is_(None)).all()
    proj_map = {str(p.id): p.name for p in projects}

    wb = Workbook()
    ws = wb.active
    ws.title = "תקציב כלל פרויקטי"
    ws.sheet_view.rightToLeft = True
    _header(ws, 1, ["פרויקט", "קטגוריה", "תיאור", "ספק", "סכום", "סוג", "תאריך"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14

    entries = db.query(models.BudgetEntry).filter(
        models.BudgetEntry.tenant_id == tenant_id,
        models.BudgetEntry.deleted_at.is_(None),
    ).order_by(models.BudgetEntry.project_id).all()

    for e in entries:
        ws.append([
            proj_map.get(str(e.project_id), ""),
            e.category,
            e.description,
            e.vendor or "",
            e.amount,
            "מתוכנן" if e.is_planned else "בפועל",
            _fmt_dt(e.entry_date),
        ])

    # הצעות מחיר
    ws2 = wb.create_sheet("הצעות מחיר")
    ws2.sheet_view.rightToLeft = True
    _header(ws2, 1, ["פרויקט", "כותרת", "ספק", "סכום", "סטטוס", "תאריך"])
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 14

    quotes = db.query(models.Quote).filter(
        models.Quote.tenant_id == tenant_id,
        models.Quote.deleted_at.is_(None),
    ).all()
    STATUS_HE = {"pending_review": "ממתין לאישור", "approved": "מאושר", "rejected": "נדחה"}
    for q in quotes:
        ws2.append([
            proj_map.get(str(q.project_id), "—") if q.project_id else "—",
            q.title,
            q.vendor or "",
            q.total_amount or 0,
            STATUS_HE.get(q.status, q.status),
            _fmt_dt(q.created_at),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="maslul_budget.xlsx"'},
    )
