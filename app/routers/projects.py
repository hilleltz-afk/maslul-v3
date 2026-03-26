import io
import os
import zipfile
from datetime import datetime, timezone
from typing import Optional
from urllib.parse import quote
from uuid import UUID

import httpx
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import crud, models, schemas
from ..deps import get_current_user_id, get_db

router = APIRouter(prefix="/tenants/{tenant_id}/projects", tags=["projects"])


def _get_project_or_404(db: Session, tenant_id: UUID, project_id: UUID) -> models.Project:
    project = (
        db.query(models.Project)
        .filter(
            models.Project.id == project_id,
            models.Project.tenant_id == tenant_id,
            models.Project.deleted_at.is_(None),
        )
        .first()
    )
    if not project:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="פרויקט לא נמצא")
    return project


@router.post("/", response_model=schemas.ProjectRead, status_code=status.HTTP_201_CREATED)
def create_project(
    tenant_id: UUID,
    project_in: schemas.ProjectCreate,
    db: Session = Depends(get_db),
    user_id: str | None = Depends(get_current_user_id),
):
    project = crud.create_entity(
        db, models.Project, project_in.model_dump(), tenant_id=str(tenant_id), created_by=user_id
    )
    db.commit()
    db.refresh(project)
    return project


@router.get("/{project_id}/export")
def export_project(
    tenant_id: UUID,
    project_id: UUID,
    db: Session = Depends(get_db),
):
    """Export project as ZIP: Excel (tasks + budget) + documents."""
    project = _get_project_or_404(db, tenant_id, project_id)

    tasks = (
        db.query(models.Task)
        .filter(models.Task.project_id == str(project_id), models.Task.deleted_at.is_(None))
        .all()
    )
    stages = (
        db.query(models.Stage)
        .filter(models.Stage.project_id == str(project_id), models.Stage.deleted_at.is_(None))
        .all()
    )
    stage_map = {str(s.id): s.name for s in stages}

    user_ids = list({t.assignee_id for t in tasks if t.assignee_id})
    users = db.query(models.User).filter(models.User.id.in_(user_ids)).all() if user_ids else []
    user_map = {str(u.id): u.name for u in users}

    budget_entries = (
        db.query(models.BudgetEntry)
        .filter(models.BudgetEntry.project_id == str(project_id), models.BudgetEntry.deleted_at.is_(None))
        .all()
    )
    documents = (
        db.query(models.Document)
        .filter(models.Document.project_id == str(project_id), models.Document.deleted_at.is_(None))
        .all()
    )

    # --- Build Excel workbook ---
    wb = openpyxl.Workbook()

    # Tasks sheet
    ws_tasks = wb.active
    ws_tasks.title = "משימות"
    task_headers = ["שם משימה", "קבוצה", "סטטוס", "עדיפות", "אחראי", "תאריך התחלה", "תאריך סיום", "תיאור"]
    ws_tasks.append(task_headers)
    STATUS_HE = {"in_progress": "בעבודה", "done": "בוצע", "delayed": "בעיכוב", "rejected": "נדחה", "partial": "בוצע חלקית"}
    PRIORITY_HE = {"high": "גבוהה", "medium": "בינונית", "low": "נמוכה"}
    for t in tasks:
        ws_tasks.append([
            t.title,
            stage_map.get(str(t.stage_id), ""),
            STATUS_HE.get(t.status, t.status),
            PRIORITY_HE.get(t.priority, t.priority),
            user_map.get(str(t.assignee_id), "") if t.assignee_id else "",
            t.start_date.strftime("%d/%m/%Y") if t.start_date else "",
            t.end_date.strftime("%d/%m/%Y") if t.end_date else "",
            t.description or "",
        ])

    # Budget sheet
    ws_budget = wb.create_sheet("תקציב")
    budget_headers = ["קטגוריה", "תיאור", "ספק", "סכום", "מתוכנן/בפועל", "הערות"]
    ws_budget.append(budget_headers)
    for e in budget_entries:
        ws_budget.append([
            e.category,
            e.description,
            e.vendor or "",
            e.amount,
            "מתוכנן" if e.is_planned else "בפועל",
            e.notes or "",
        ])

    # Documents sheet
    ws_docs = wb.create_sheet("מסמכים")
    ws_docs.append(["שם מסמך", "קישור", "תאריך תוקף"])
    for d in documents:
        ws_docs.append([
            d.name,
            d.path,
            d.expiry_date.strftime("%d/%m/%Y") if d.expiry_date else "",
        ])

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    # --- Build ZIP ---
    zip_buf = io.BytesIO()
    project_name_safe = project.name.replace("/", "-").replace("\\", "-")

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{project_name_safe}/נתונים.xlsx", excel_buf.read())

        api_base = os.getenv("INTERNAL_API_BASE", "http://localhost:8000")
        for doc in documents:
            try:
                if doc.path.startswith("http"):
                    resp = httpx.get(doc.path, timeout=15)
                    file_bytes = resp.content
                else:
                    local_path = doc.path.lstrip("/")
                    full_path = os.path.join(os.getcwd(), local_path)
                    with open(full_path, "rb") as f:
                        file_bytes = f.read()
                ext = os.path.splitext(doc.path)[1] or ""
                safe_name = doc.name.replace("/", "-")
                if not safe_name.endswith(ext):
                    safe_name += ext
                zf.writestr(f"{project_name_safe}/מסמכים/{safe_name}", file_bytes)
            except Exception:
                pass  # skip unavailable files

    zip_buf.seek(0)

    filename = f"{project_name_safe}.zip"
    filename_encoded = quote(filename, safe="")
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=\"project.zip\"; filename*=UTF-8''{filename_encoded}"},
    )


@router.get("/{project_id}", response_model=schemas.ProjectRead)
def read_project(tenant_id: UUID, project_id: UUID, db: Session = Depends(get_db)):
    return _get_project_or_404(db, tenant_id, project_id)


@router.get("/", response_model=list[schemas.ProjectRead])
def list_projects(
    tenant_id: UUID,
    archived: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(models.Project).filter(
        models.Project.tenant_id == tenant_id,
        models.Project.deleted_at.is_(None),
    )
    if archived is True:
        q = q.filter(models.Project.archived_at.isnot(None))
    elif archived is False or archived is None:
        q = q.filter(models.Project.archived_at.is_(None))
    return q.all()


@router.put("/{project_id}", response_model=schemas.ProjectRead)
def update_project(
    tenant_id: UUID,
    project_id: UUID,
    project_in: schemas.ProjectUpdate,
    db: Session = Depends(get_db),
    changed_by: str | None = Depends(get_current_user_id),
):
    project = _get_project_or_404(db, tenant_id, project_id)
    project = crud.update_entity(db, project, project_in.model_dump(), changed_by=changed_by)
    db.commit()
    db.refresh(project)
    return project


@router.post("/{project_id}/archive", response_model=schemas.ProjectRead)
def archive_project(
    tenant_id: UUID,
    project_id: UUID,
    db: Session = Depends(get_db),
):
    project = _get_project_or_404(db, tenant_id, project_id)
    project.archived_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(project)
    return project


@router.delete("/{project_id}/archive", response_model=schemas.ProjectRead)
def unarchive_project(
    tenant_id: UUID,
    project_id: UUID,
    db: Session = Depends(get_db),
):
    project = _get_project_or_404(db, tenant_id, project_id)
    project.archived_at = None
    db.commit()
    db.refresh(project)
    return project


@router.delete("/{project_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_project(
    tenant_id: UUID,
    project_id: UUID,
    db: Session = Depends(get_db),
    changed_by: str | None = Depends(get_current_user_id),
):
    project = _get_project_or_404(db, tenant_id, project_id)
    now = datetime.now(timezone.utc)

    # Cascade soft-delete: tasks, stages, documents, budget entries
    for model_cls in (models.Task, models.Stage, models.Document, models.BudgetEntry):
        db.query(model_cls).filter(
            model_cls.project_id == str(project_id),
            model_cls.deleted_at.is_(None),
        ).update({"deleted_at": now}, synchronize_session=False)

    crud.soft_delete_entity(db, project, changed_by=changed_by)
    db.commit()
    return None


# ---- members sub-resource (kept here for routing clarity) ----

@router.get("/{project_id}/members/", response_model=list[schemas.ProjectMemberRead])
def list_members(tenant_id: UUID, project_id: UUID, db: Session = Depends(get_db)):
    _get_project_or_404(db, tenant_id, project_id)
    members = (
        db.query(models.ProjectMember)
        .filter(
            models.ProjectMember.project_id == str(project_id),
            models.ProjectMember.deleted_at.is_(None),
        )
        .all()
    )
    result = []
    for m in members:
        user = db.query(models.User).filter(models.User.id == m.user_id).first()
        result.append(schemas.ProjectMemberRead(
            id=m.id,
            project_id=m.project_id,
            user_id=m.user_id,
            role=m.role,
            created_at=m.created_at,
            updated_at=m.updated_at,
            user_name=user.name if user else None,
            user_email=user.email if user else None,
        ))
    return result


@router.post("/{project_id}/members/", response_model=schemas.ProjectMemberRead, status_code=status.HTTP_201_CREATED)
def add_member(
    tenant_id: UUID,
    project_id: UUID,
    member_in: schemas.ProjectMemberCreate,
    db: Session = Depends(get_db),
):
    _get_project_or_404(db, tenant_id, project_id)
    existing = (
        db.query(models.ProjectMember)
        .filter(
            models.ProjectMember.project_id == str(project_id),
            models.ProjectMember.user_id == str(member_in.user_id),
            models.ProjectMember.deleted_at.is_(None),
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="המשתמש כבר חבר בפרויקט")
    m = models.ProjectMember(
        project_id=str(project_id),
        user_id=str(member_in.user_id),
        role=member_in.role,
    )
    db.add(m)
    db.commit()
    db.refresh(m)
    user = db.query(models.User).filter(models.User.id == m.user_id).first()
    return schemas.ProjectMemberRead(
        id=m.id,
        project_id=m.project_id,
        user_id=m.user_id,
        role=m.role,
        created_at=m.created_at,
        updated_at=m.updated_at,
        user_name=user.name if user else None,
        user_email=user.email if user else None,
    )


@router.put("/{project_id}/members/{user_id}", response_model=schemas.ProjectMemberRead)
def update_member_role(
    tenant_id: UUID,
    project_id: UUID,
    user_id: UUID,
    member_in: schemas.ProjectMemberUpdate,
    db: Session = Depends(get_db),
):
    _get_project_or_404(db, tenant_id, project_id)
    m = (
        db.query(models.ProjectMember)
        .filter(
            models.ProjectMember.project_id == str(project_id),
            models.ProjectMember.user_id == str(user_id),
            models.ProjectMember.deleted_at.is_(None),
        )
        .first()
    )
    if not m:
        raise HTTPException(status_code=404, detail="חבר לא נמצא")
    m.role = member_in.role
    db.commit()
    db.refresh(m)
    user = db.query(models.User).filter(models.User.id == m.user_id).first()
    return schemas.ProjectMemberRead(
        id=m.id, project_id=m.project_id, user_id=m.user_id, role=m.role,
        created_at=m.created_at, updated_at=m.updated_at,
        user_name=user.name if user else None, user_email=user.email if user else None,
    )


@router.delete("/{project_id}/members/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_member(
    tenant_id: UUID,
    project_id: UUID,
    user_id: UUID,
    db: Session = Depends(get_db),
):
    _get_project_or_404(db, tenant_id, project_id)
    m = (
        db.query(models.ProjectMember)
        .filter(
            models.ProjectMember.project_id == str(project_id),
            models.ProjectMember.user_id == str(user_id),
            models.ProjectMember.deleted_at.is_(None),
        )
        .first()
    )
    if not m:
        raise HTTPException(status_code=404, detail="חבר לא נמצא")
    m.deleted_at = datetime.now(timezone.utc)
    db.commit()
    return None
